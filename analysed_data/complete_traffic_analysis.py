import glob
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "analysed_data"
PLOTS_DIR = OUTPUT_DIR / "plots"
RAW_DIR = ROOT / "raw traffic data"
RESULT_DIR = ROOT / "result"
START_DATE = pd.Timestamp("2026-05-25").date()
END_DATE = pd.Timestamp("2026-05-31").date()
DEFAULT_DISTANCE_KM = 10.0


def progress(message):
    print(f"\n{message}")
    print("-" * len(message))


def warn(message):
    print(f"WARNING: {message}")


def ensure_optional_plot_packages():
    if os.environ.get("TRAFFIC_ANALYSIS_SKIP_PIP") == "1":
        print("Skipping automatic pip install because TRAFFIC_ANALYSIS_SKIP_PIP=1.")
        return False

    missing = []
    for package in ["matplotlib", "seaborn"]:
        try:
            __import__(package)
        except ImportError:
            missing.append(package)
    if not missing:
        return True

    print(f"Missing optional plotting packages: {', '.join(missing)}")
    print("Trying automatic pip install. If this fails, Pillow fallback plots will be used.")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
    except Exception as exc:
        warn(f"Could not install {missing}: {exc}")
        return False

    try:
        for package in missing:
            __import__(package)
        return True
    except ImportError:
        return False


def safe_read_csv(path):
    try:
        return pd.read_csv(path)
    except Exception as exc:
        warn(f"Skipping unreadable file {path}: {exc}")
        return None


def clean_col(value):
    return str(value).strip().lower().replace(" ", "_")


def detect_column(columns, semantic_name):
    normalized = {clean_col(col): col for col in columns}
    rules = {
        "junction": [
            ["junction_name"],
            ["junction", "name"],
            ["junction"],
            ["source_row"],
        ],
        "hour": [["hour"], ["requested_departure_local"], ["observed_at_local"]],
        "date": [["date"], ["requested_departure_local"], ["observed_at_local"]],
        "day_of_week": [["day_of_week"], ["weekday"]],
        "duration_s": [
            ["avg_traffic_duration_s"],
            ["traffic_duration_s"],
            ["duration_s"],
            ["duration"],
        ],
        "congestion_index": [["avg_congestion_index"], ["congestion_index"]],
        "speed_kmph": [["avg_traffic_speed_kmh"], ["traffic_speed_kmh"], ["speed"]],
        "distance_m": [["route_distance_m"], ["distance_m"], ["distance"]],
        "api_status": [["api_status"], ["status"]],
    }
    for keywords in rules.get(semantic_name, []):
        for norm, original in normalized.items():
            if all(keyword in norm for keyword in keywords):
                return original
    return None


def to_hour(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return int(value)
    text = str(value).strip()
    if ":" in text and len(text) <= 8:
        try:
            return int(text.split(":")[0])
        except ValueError:
            return np.nan
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.notna(parsed):
        return int(parsed.hour)
    try:
        return int(float(text))
    except ValueError:
        return np.nan


def to_date(value):
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return pd.NaT
    return parsed.date()


def date_in_range(value):
    if pd.isna(value):
        return False
    return START_DATE <= value <= END_DATE


def in_date_range_df(df, date_col):
    dates = df[date_col].apply(to_date)
    return df.loc[dates.apply(date_in_range)].copy()


def file_inventory():
    files = []
    for folder in [RAW_DIR, RESULT_DIR]:
        if folder.exists():
            files.extend(path for path in folder.rglob("*") if path.is_file())
    return sorted(files)


def explore_files(files):
    records = []
    for path in files:
        if path.suffix.lower() != ".csv":
            continue
        df = safe_read_csv(path)
        if df is None:
            continue
        detected = {
            key: detect_column(df.columns, key)
            for key in [
                "junction",
                "hour",
                "date",
                "day_of_week",
                "duration_s",
                "congestion_index",
                "speed_kmph",
                "distance_m",
                "api_status",
            ]
        }
        print(f"{path.name}: {df.shape[0]} rows x {df.shape[1]} columns")
        print(f"  columns: {list(df.columns)}")
        print(f"  detected: {detected}")
        records.append(
            {
                "file": str(path.relative_to(ROOT)),
                "rows": df.shape[0],
                "columns": df.shape[1],
                "column_names": ", ".join(map(str, df.columns)),
                **{f"detected_{key}": value for key, value in detected.items()},
            }
        )
    out = pd.DataFrame(records)
    out.to_csv(OUTPUT_DIR / "data_exploration_summary.csv", index=False)
    return out


def parse_date_from_file(path):
    text = path.name
    match = pd.Series([text]).str.extract(r"(2026-\d{2}-\d{2})")[0].iloc[0]
    if pd.notna(match):
        return pd.Timestamp(match).date()
    return None


def load_daily_hourly_data():
    files = sorted(RESULT_DIR.glob("2026-05-*_hourly_data/junction_hourly_traffic_summary_*.csv"))
    rows = []
    for path in files:
        file_date = parse_date_from_file(path)
        if file_date is not None and not date_in_range(file_date):
            continue
        df = safe_read_csv(path)
        if df is None or df.empty:
            continue
        cols = df.columns
        mapping = {
            "junction_id": detect_column(cols, "junction"),
            "date": detect_column(cols, "date"),
            "hour": detect_column(cols, "hour"),
            "duration_s": detect_column(cols, "duration_s"),
            "congestion_index": detect_column(cols, "congestion_index"),
            "speed_kmph": detect_column(cols, "speed_kmph"),
        }
        required = ["junction_id", "date", "hour", "duration_s", "speed_kmph"]
        missing = [key for key in required if mapping[key] is None]
        if missing:
            warn(f"Skipping {path.name}; missing detected columns: {missing}")
            continue
        standardized = pd.DataFrame()
        standardized["model"] = df[mapping["model"]] if mapping.get("model") else df.get("model", "UNKNOWN")
        standardized["date"] = df[mapping["date"]].apply(to_date)
        standardized["day_of_week"] = pd.to_datetime(standardized["date"]).dt.day_name()
        standardized["junction_id"] = df[mapping["junction_id"]].astype(str)
        standardized["source_row"] = df.get("source_row", "")
        standardized["junction_latitude"] = df.get("junction_latitude", np.nan)
        standardized["junction_longitude"] = df.get("junction_longitude", np.nan)
        standardized["hour"] = df[mapping["hour"]].apply(to_hour).astype("Int64")
        standardized["duration_s"] = pd.to_numeric(df[mapping["duration_s"]], errors="coerce")
        standardized["google_speed_kmph"] = pd.to_numeric(df[mapping["speed_kmph"]], errors="coerce")
        standardized["congestion_index"] = (
            pd.to_numeric(df[mapping["congestion_index"]], errors="coerce")
            if mapping["congestion_index"]
            else np.nan
        )
        static_col = detect_column(cols, "static_duration_s") or "avg_static_duration_s"
        standardized["static_duration_s"] = (
            pd.to_numeric(df[static_col], errors="coerce") if static_col in df.columns else np.nan
        )
        delay_col = "avg_delay_s" if "avg_delay_s" in df.columns else detect_column(cols, "delay_s")
        standardized["google_static_delay_s"] = (
            pd.to_numeric(df[delay_col], errors="coerce") if delay_col in df.columns else np.nan
        )
        standardized["records_from_summary"] = pd.to_numeric(df.get("records", 1), errors="coerce")
        standardized = standardized[standardized["date"].apply(date_in_range)]
        rows.append(standardized)
    if not rows:
        raise RuntimeError("No daily junction hourly files were loaded.")
    master = pd.concat(rows, ignore_index=True)
    master = master.dropna(subset=["date", "junction_id", "hour", "duration_s"])
    master["hour"] = master["hour"].astype(int)
    return master


def load_peak_data():
    files = sorted(RESULT_DIR.glob("2026-05-*_hourly_data/junction_peak_hour_summary_*.csv"))
    rows = []
    for path in files:
        df = safe_read_csv(path)
        if df is None or df.empty:
            continue
        date_col = detect_column(df.columns, "date")
        junction_col = detect_column(df.columns, "junction")
        if date_col is None or junction_col is None:
            warn(f"Skipping peak file {path.name}; missing date or junction column.")
            continue
        df = in_date_range_df(df, date_col)
        if df.empty:
            continue
        df["date"] = df[date_col].apply(to_date)
        df["day_of_week"] = pd.to_datetime(df["date"]).dt.day_name()
        df["junction_id"] = df[junction_col].astype(str)
        rows.append(df)
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def load_weekly_peak_crosscheck():
    files = sorted(RESULT_DIR.glob("2026-05-25_to_2026-05-31_weekly_data/junction_weekly_peak_hour_summary_*.csv"))
    rows = []
    for path in files:
        df = safe_read_csv(path)
        if df is not None and not df.empty:
            junction_col = detect_column(df.columns, "junction")
            if junction_col:
                df["junction_id"] = df[junction_col].astype(str)
                rows.append(df)
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def load_raw_data():
    files = sorted(RAW_DIR.glob("*traffic_observations_*.csv"))
    rows = []
    for path in files:
        df = safe_read_csv(path)
        if df is None or df.empty:
            continue
        date_col = detect_column(df.columns, "date")
        hour_col = detect_column(df.columns, "hour")
        junction_col = detect_column(df.columns, "junction")
        if date_col is None or hour_col is None or junction_col is None:
            warn(f"Skipping raw file {path.name}; date/hour/junction columns not detected.")
            continue
        df["date"] = df[date_col].apply(to_date)
        df = df[df["date"].apply(date_in_range)].copy()
        if df.empty:
            continue
        df["day_of_week"] = pd.to_datetime(df["date"]).dt.day_name()
        df["hour"] = df[hour_col].apply(to_hour)
        df["junction_id"] = df[junction_col].astype(str)
        rows.append(df)
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def distance_by_junction(raw):
    if raw.empty:
        return pd.DataFrame(columns=["junction_id", "distance_km_used", "distance_source"])
    distance_col = detect_column(raw.columns, "distance_m")
    if distance_col is None:
        return pd.DataFrame(columns=["junction_id", "distance_km_used", "distance_source"])
    raw = raw.copy()
    raw[distance_col] = pd.to_numeric(raw[distance_col], errors="coerce")
    dist = (
        raw.dropna(subset=[distance_col])
        .groupby("junction_id", as_index=False)[distance_col]
        .mean()
    )
    dist["distance_km_used"] = dist[distance_col] / 1000.0
    dist["distance_source"] = "Google route_distance_m"
    return dist[["junction_id", "distance_km_used", "distance_source"]]


def add_distance(master, raw):
    dist = distance_by_junction(raw)
    enriched = master.merge(dist, on="junction_id", how="left")
    enriched["distance_km_used"] = enriched["distance_km_used"].fillna(DEFAULT_DISTANCE_KM)
    enriched["distance_source"] = enriched["distance_source"].fillna("Default 10 km")
    return enriched


def basic_stats(master):
    grouped = master.groupby("junction_id")
    stats = grouped.agg(
        min_travel_duration_s=("duration_s", "min"),
        max_travel_duration_s=("duration_s", "max"),
        mean_travel_duration_s=("duration_s", "mean"),
        min_speed_kmph=("google_speed_kmph", "min"),
        max_speed_kmph=("google_speed_kmph", "max"),
        mean_speed_kmph=("google_speed_kmph", "mean"),
        mean_congestion_index=("congestion_index", "mean"),
        total_hours_data_available=("duration_s", "count"),
        junction_latitude=("junction_latitude", "first"),
        junction_longitude=("junction_longitude", "first"),
    ).reset_index()
    return stats.round(4)


def free_flow_speed(master):
    offpeak = master[master["hour"].between(0, 5)]
    min_duration = master.groupby("junction_id", as_index=False).agg(
        free_flow_min_duration_s=("duration_s", "min"),
        distance_km_used=("distance_km_used", "first"),
        distance_source=("distance_source", "first"),
    )
    min_duration["method1_free_flow_speed_kmph"] = (
        min_duration["distance_km_used"] / (min_duration["free_flow_min_duration_s"] / 3600.0)
    )
    method2 = (
        offpeak.groupby("junction_id", as_index=False)["google_speed_kmph"]
        .mean()
        .rename(columns={"google_speed_kmph": "method2_offpeak_avg_speed_kmph"})
    )
    result = min_duration.merge(method2, on="junction_id", how="left")
    result["primary_free_flow_speed_kmph"] = result["method1_free_flow_speed_kmph"]
    return result.round(4)


def delay_data(master, freeflow):
    ff = freeflow[["junction_id", "free_flow_min_duration_s", "primary_free_flow_speed_kmph"]]
    df = master.merge(ff, on="junction_id", how="left")
    df["delay_seconds"] = df["duration_s"] - df["free_flow_min_duration_s"]
    df["delay_minutes"] = df["delay_seconds"] / 60.0
    conditions = [
        df["delay_seconds"] <= 0,
        (df["delay_seconds"] > 0) & (df["delay_seconds"] <= 60),
        (df["delay_seconds"] > 60) & (df["delay_seconds"] <= 300),
        df["delay_seconds"] > 300,
    ]
    labels = ["Free Flow", "Minor Delay", "Moderate Delay", "Severe Delay"]
    df["delay_flag"] = np.select(conditions, labels, default="Unknown")
    return df


def daily_average_delay(delay_df):
    return (
        delay_df.groupby(["junction_id", "date", "day_of_week"], as_index=False)
        .agg(
            avg_delay_seconds=("delay_seconds", "mean"),
            avg_delay_minutes=("delay_minutes", "mean"),
            max_delay_minutes=("delay_minutes", "max"),
            avg_duration_s=("duration_s", "mean"),
            avg_speed_kmph=("google_speed_kmph", "mean"),
            avg_congestion_index=("congestion_index", "mean"),
            records=("duration_s", "count"),
        )
        .round(4)
    )


def hourly_delay_profile(delay_df):
    return (
        delay_df.groupby(["junction_id", "hour"], as_index=False)
        .agg(
            avg_delay_seconds=("delay_seconds", "mean"),
            avg_delay_minutes=("delay_minutes", "mean"),
            avg_duration_s=("duration_s", "mean"),
            avg_speed_kmph=("google_speed_kmph", "mean"),
            avg_congestion_index=("congestion_index", "mean"),
            records=("duration_s", "count"),
        )
        .round(4)
    )


def velocity_profiles(delay_df):
    df = delay_df.copy()
    df["calculated_speed_kmph"] = df["distance_km_used"] / (df["duration_s"] / 3600.0)
    df["google_speed_difference_kmph"] = df["calculated_speed_kmph"] - df["google_speed_kmph"]
    profile = (
        df.groupby(["junction_id", "hour"], as_index=False)
        .agg(
            avg_calculated_speed_kmph=("calculated_speed_kmph", "mean"),
            avg_google_speed_kmph=("google_speed_kmph", "mean"),
            avg_speed_difference_kmph=("google_speed_difference_kmph", "mean"),
            free_flow_speed_kmph=("primary_free_flow_speed_kmph", "first"),
            avg_duration_s=("duration_s", "mean"),
            records=("duration_s", "count"),
        )
        .round(4)
    )
    return df, profile


def color_palette(n):
    base = [
        (31, 119, 180),
        (255, 127, 14),
        (44, 160, 44),
        (214, 39, 40),
        (148, 103, 189),
        (140, 86, 75),
        (227, 119, 194),
        (127, 127, 127),
        (188, 189, 34),
        (23, 190, 207),
    ]
    return [base[i % len(base)] for i in range(n)]


def font(size=14):
    try:
        return ImageFont.truetype("arial.ttf", size)
    except OSError:
        return ImageFont.load_default()


def nice_range(values):
    values = [float(v) for v in values if pd.notna(v) and np.isfinite(v)]
    if not values:
        return 0, 1
    lo, hi = min(values), max(values)
    if lo == hi:
        return lo - 1, hi + 1
    pad = (hi - lo) * 0.08
    return lo - pad, hi + pad


def draw_line_plot(series, output, title, x_label, y_label, y_min=None, y_max=None, hlines=None):
    width, height = 1400, 850
    margin_left, margin_right = 95, 260
    margin_top, margin_bottom = 75, 90
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    title_font, label_font, small_font = font(24), font(15), font(12)
    plot_w = width - margin_left - margin_right
    plot_h = height - margin_top - margin_bottom
    all_y = []
    for item in series:
        all_y.extend(item["y"])
    if hlines:
        all_y.extend([line["y"] for line in hlines])
    ymin, ymax = nice_range(all_y)
    if y_min is not None:
        ymin = y_min
    if y_max is not None:
        ymax = y_max
    xmin, xmax = 0, 23

    def sx(x):
        return margin_left + ((x - xmin) / (xmax - xmin)) * plot_w

    def sy(y):
        if ymax == ymin:
            return margin_top + plot_h / 2
        return margin_top + plot_h - ((y - ymin) / (ymax - ymin)) * plot_h

    draw.text((margin_left, 24), title, fill="black", font=title_font)
    draw.line((margin_left, margin_top, margin_left, margin_top + plot_h), fill="black", width=2)
    draw.line((margin_left, margin_top + plot_h, margin_left + plot_w, margin_top + plot_h), fill="black", width=2)
    for h in range(0, 24, 2):
        x = sx(h)
        draw.line((x, margin_top, x, margin_top + plot_h), fill=(235, 235, 235))
        draw.text((x - 10, margin_top + plot_h + 10), str(h), fill="black", font=small_font)
    for i in range(6):
        yv = ymin + (ymax - ymin) * i / 5
        y = sy(yv)
        draw.line((margin_left, y, margin_left + plot_w, y), fill=(235, 235, 235))
        draw.text((8, y - 8), f"{yv:.1f}", fill="black", font=small_font)
    draw.text((margin_left + plot_w / 2 - 30, height - 40), x_label, fill="black", font=label_font)
    draw.text((10, margin_top - 25), y_label, fill="black", font=label_font)

    if hlines:
        for line in hlines:
            y = sy(line["y"])
            for x in range(int(margin_left), int(margin_left + plot_w), 18):
                draw.line((x, y, x + 9, y), fill=line.get("color", (210, 0, 0)), width=2)
            draw.text((margin_left + plot_w + 15, y - 8), line["label"], fill=line.get("color", (210, 0, 0)), font=small_font)

    for item in series:
        points = [(sx(x), sy(y)) for x, y in zip(item["x"], item["y"]) if pd.notna(y)]
        if len(points) >= 2:
            draw.line(points, fill=item["color"], width=item.get("width", 2))
        for x, y in points:
            draw.ellipse((x - 2, y - 2, x + 2, y + 2), fill=item["color"])
    legend_x = margin_left + plot_w + 15
    legend_y = margin_top
    for i, item in enumerate(series[:30]):
        y = legend_y + i * 22
        draw.line((legend_x, y + 8, legend_x + 25, y + 8), fill=item["color"], width=3)
        draw.text((legend_x + 32, y), str(item["label"])[:24], fill="black", font=small_font)
    img.save(output)


def draw_bar_chart(labels, morning, evening, output, title):
    width, height = 1500, 850
    margin_left, margin_right = 90, 40
    margin_top, margin_bottom = 80, 160
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    title_font, small_font = font(24), font(11)
    plot_w = width - margin_left - margin_right
    plot_h = height - margin_top - margin_bottom
    all_values = list(morning) + list(evening)
    ymin = min([0] + all_values)
    ymax = max([0] + all_values)
    if ymin == ymax:
        ymin, ymax = -1, 1
    else:
        pad = (ymax - ymin) * 0.12
        ymin -= pad
        ymax += pad

    def sy(value):
        return margin_top + plot_h - ((value - ymin) / (ymax - ymin)) * plot_h

    baseline_y = sy(0)
    draw.text((margin_left, 25), title, fill="black", font=title_font)
    draw.line((margin_left, margin_top, margin_left, margin_top + plot_h), fill="black", width=2)
    draw.line((margin_left, baseline_y, margin_left + plot_w, baseline_y), fill="black", width=2)
    n = len(labels)
    band = plot_w / max(n, 1)
    bar_w = band * 0.32
    for i in range(6):
        value = ymin + (ymax - ymin) * i / 5
        y = sy(value)
        draw.line((margin_left, y, margin_left + plot_w, y), fill=(235, 235, 235))
        draw.text((8, y - 8), f"{value:.0f}%", fill="black", font=small_font)
    for i, label in enumerate(labels):
        cx = margin_left + i * band + band / 2
        for offset, value, color in [(-bar_w / 2, morning[i], (56, 112, 164)), (bar_w / 2, evening[i], (198, 87, 87))]:
            x0 = cx + offset - bar_w / 2
            x1 = cx + offset + bar_w / 2
            y_value = sy(value)
            y0, y1 = sorted([y_value, baseline_y])
            draw.rectangle((x0, y0, x1, y1), fill=color)
        draw.text((cx - 12, margin_top + plot_h + 8), str(label), fill="black", font=small_font)
    draw.rectangle((margin_left + 20, margin_top + 10, margin_left + 35, margin_top + 25), fill=(56, 112, 164))
    draw.text((margin_left + 42, margin_top + 8), "Morning", fill="black", font=small_font)
    draw.rectangle((margin_left + 120, margin_top + 10, margin_left + 135, margin_top + 25), fill=(198, 87, 87))
    draw.text((margin_left + 142, margin_top + 8), "Evening", fill="black", font=small_font)
    img.save(output)


def draw_scatter_plot(x_values, y_values, output, title, x_label, y_label):
    width, height = 1200, 800
    margin_left, margin_right = 90, 40
    margin_top, margin_bottom = 75, 80
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    title_font, label_font, small_font = font(24), font(14), font(11)
    x = pd.Series(x_values).dropna()
    y = pd.Series(y_values).dropna()
    paired = pd.DataFrame({"x": x_values, "y": y_values}).dropna()
    xmin, xmax = nice_range(paired["x"])
    ymin, ymax = nice_range(paired["y"])
    plot_w = width - margin_left - margin_right
    plot_h = height - margin_top - margin_bottom

    def sx(v):
        return margin_left + ((v - xmin) / (xmax - xmin)) * plot_w

    def sy(v):
        return margin_top + plot_h - ((v - ymin) / (ymax - ymin)) * plot_h

    draw.text((margin_left, 25), title, fill="black", font=title_font)
    draw.line((margin_left, margin_top, margin_left, margin_top + plot_h), fill="black", width=2)
    draw.line((margin_left, margin_top + plot_h, margin_left + plot_w, margin_top + plot_h), fill="black", width=2)
    for i in range(6):
        xv = xmin + (xmax - xmin) * i / 5
        yv = ymin + (ymax - ymin) * i / 5
        xp = sx(xv)
        yp = sy(yv)
        draw.line((xp, margin_top, xp, margin_top + plot_h), fill=(235, 235, 235))
        draw.line((margin_left, yp, margin_left + plot_w, yp), fill=(235, 235, 235))
        draw.text((xp - 18, margin_top + plot_h + 8), f"{xv:.2f}", fill="black", font=small_font)
        draw.text((8, yp - 8), f"{yv:.1f}", fill="black", font=small_font)
    for _, row in paired.iterrows():
        px, py = sx(row["x"]), sy(row["y"])
        draw.ellipse((px - 2, py - 2, px + 2, py + 2), fill=(40, 105, 180))
    draw.text((margin_left + plot_w / 2 - 80, height - 38), x_label, fill="black", font=label_font)
    draw.text((8, margin_top - 25), y_label, fill="black", font=label_font)
    img.save(output)


def draw_heatmap(pivot, output, title):
    data = pivot.copy()
    values = data.to_numpy(dtype=float)
    vmin = np.nanmin(values)
    vmax = np.nanmax(values)
    rows, cols = values.shape
    cell_w, cell_h = 42, 24
    left, top = 110, 75
    width = left + cols * cell_w + 90
    height = top + rows * cell_h + 80
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    title_font, small_font = font(22), font(10)
    draw.text((left, 25), title, fill="black", font=title_font)
    for j, col in enumerate(data.columns):
        draw.text((left + j * cell_w + 8, top - 22), str(col), fill="black", font=small_font)
    for i, idx in enumerate(data.index):
        draw.text((8, top + i * cell_h + 6), str(idx), fill="black", font=small_font)
        for j in range(cols):
            value = values[i, j]
            if np.isnan(value) or vmax == vmin:
                color = (230, 230, 230)
            else:
                ratio = (value - vmin) / (vmax - vmin)
                color = (int(80 + 175 * ratio), int(210 - 150 * ratio), 70)
            x0 = left + j * cell_w
            y0 = top + i * cell_h
            draw.rectangle((x0, y0, x0 + cell_w - 1, y0 + cell_h - 1), fill=color)
    draw.text((left, height - 45), f"Green = lower duration, Red = higher duration. Range {vmin:.1f}s to {vmax:.1f}s", fill="black", font=small_font)
    img.save(output)


def peak_hours(delay_df, peak_crosscheck, weekly_crosscheck):
    avg_by_hour = (
        delay_df.groupby(["junction_id", "hour"], as_index=False)
        .agg(
            avg_duration_s=("duration_s", "mean"),
            avg_speed_kmph=("google_speed_kmph", "mean"),
            avg_delay_minutes=("delay_minutes", "mean"),
            avg_congestion_index=("congestion_index", "mean"),
            records=("duration_s", "count"),
        )
    )
    rows = []
    for junction, group in avg_by_hour.groupby("junction_id"):
        morning = group[group["hour"].between(7, 10)]
        evening = group[group["hour"].between(16, 20)]
        morning_peak = morning.loc[morning["avg_duration_s"].idxmax()] if not morning.empty else pd.Series()
        evening_peak = evening.loc[evening["avg_duration_s"].idxmax()] if not evening.empty else pd.Series()
        raw_j = delay_df[delay_df["junction_id"] == junction]
        worst = raw_j.loc[raw_j["duration_s"].idxmax()]
        best = raw_j.loc[raw_j["duration_s"].idxmin()]
        daily_cross = ""
        if not peak_crosscheck.empty and "peak_hour" in peak_crosscheck.columns:
            daily_hours = (
                peak_crosscheck[peak_crosscheck["junction_id"] == junction]["peak_hour"]
                .astype(str)
                .value_counts()
            )
            daily_cross = daily_hours.index[0] if not daily_hours.empty else ""
        weekly_cross = ""
        if not weekly_crosscheck.empty and "peak_hour" in weekly_crosscheck.columns:
            vals = weekly_crosscheck[weekly_crosscheck["junction_id"] == junction]["peak_hour"].astype(str).unique()
            weekly_cross = vals[0] if len(vals) else ""
        rows.append(
            {
                "junction_id": junction,
                "morning_peak_hour": int(morning_peak.get("hour", -1)) if not morning_peak.empty else "",
                "morning_peak_avg_duration_s": morning_peak.get("avg_duration_s", np.nan),
                "evening_peak_hour": int(evening_peak.get("hour", -1)) if not evening_peak.empty else "",
                "evening_peak_avg_duration_s": evening_peak.get("avg_duration_s", np.nan),
                "worst_hour_overall": int(worst["hour"]),
                "worst_date": worst["date"],
                "worst_duration_s": worst["duration_s"],
                "best_hour_overall": int(best["hour"]),
                "best_date": best["date"],
                "best_duration_s": best["duration_s"],
                "daily_peak_crosscheck_mode_hour": daily_cross,
                "weekly_peak_crosscheck_hour": weekly_cross,
            }
        )
    return pd.DataFrame(rows).round(4)


def speed_reduction(delay_df, peak_df, freeflow):
    rows = []
    for _, peak in peak_df.iterrows():
        junction = peak["junction_id"]
        ff_speed = freeflow.loc[
            freeflow["junction_id"] == junction,
            "primary_free_flow_speed_kmph",
        ].iloc[0]
        jdf = delay_df[delay_df["junction_id"] == junction]
        morning_hour = peak["morning_peak_hour"]
        evening_hour = peak["evening_peak_hour"]
        morning_speed = jdf[jdf["hour"] == morning_hour]["google_speed_kmph"].mean()
        evening_speed = jdf[jdf["hour"] == evening_hour]["google_speed_kmph"].mean()
        morning_delay = jdf[jdf["hour"] == morning_hour]["delay_minutes"].mean()
        evening_delay = jdf[jdf["hour"] == evening_hour]["delay_minutes"].mean()
        morning_red = ff_speed - morning_speed
        evening_red = ff_speed - evening_speed
        rows.append(
            {
                "junction_id": junction,
                "free_flow_speed_kmph": ff_speed,
                "morning_peak_hour": morning_hour,
                "morning_peak_speed_kmph": morning_speed,
                "morning_peak_delay_minutes": morning_delay,
                "morning_speed_reduction_kmph": morning_red,
                "morning_speed_reduction_pct": (morning_red / ff_speed) * 100 if ff_speed else np.nan,
                "evening_peak_hour": evening_hour,
                "evening_peak_speed_kmph": evening_speed,
                "evening_peak_delay_minutes": evening_delay,
                "evening_speed_reduction_kmph": evening_red,
                "evening_speed_reduction_pct": (evening_red / ff_speed) * 100 if ff_speed else np.nan,
            }
        )
    result = pd.DataFrame(rows)
    result["max_peak_speed_reduction_pct"] = result[
        ["morning_speed_reduction_pct", "evening_speed_reduction_pct"]
    ].max(axis=1)
    result = result.sort_values("max_peak_speed_reduction_pct", ascending=False).reset_index(drop=True)
    result["congestion_rank"] = np.arange(1, len(result) + 1)
    return result.round(4)


def congestion_analysis(delay_df):
    hourly = (
        delay_df.groupby(["junction_id", "hour"], as_index=False)
        .agg(
            avg_congestion_index=("congestion_index", "mean"),
            avg_delay_minutes=("delay_minutes", "mean"),
            avg_duration_s=("duration_s", "mean"),
        )
    )
    rows = []
    corr_df = delay_df[["congestion_index", "delay_minutes"]].dropna()
    corr = corr_df["congestion_index"].corr(corr_df["delay_minutes"]) if len(corr_df) >= 2 else np.nan
    for junction, group in hourly.groupby("junction_id"):
        peak = group.loc[group["avg_congestion_index"].idxmax()]
        rows.append(
            {
                "junction_id": junction,
                "max_congestion_hour": int(peak["hour"]),
                "max_avg_congestion_index": peak["avg_congestion_index"],
                "avg_delay_at_max_congestion_minutes": peak["avg_delay_minutes"],
                "mean_congestion_index": group["avg_congestion_index"].mean(),
                "mean_delay_minutes": group["avg_delay_minutes"].mean(),
                "overall_correlation_congestion_delay": corr,
            }
        )
    return pd.DataFrame(rows).round(4), hourly.round(4), corr


def daywise_comparison(delay_df):
    return (
        delay_df.groupby(["date", "day_of_week"], as_index=False)
        .agg(
            avg_speed_kmph=("google_speed_kmph", "mean"),
            avg_delay_minutes=("delay_minutes", "mean"),
            avg_duration_s=("duration_s", "mean"),
            avg_congestion_index=("congestion_index", "mean"),
            records=("duration_s", "count"),
        )
        .round(4)
    )


def make_plots(delay_df, velocity_profile, speed_reduction_df, congestion_hourly):
    PLOTS_DIR.mkdir(parents=True, exist_ok=True)
    junctions = sorted(velocity_profile["junction_id"].unique(), key=lambda x: float(x) if str(x).replace(".", "", 1).isdigit() else str(x))
    colors = color_palette(len(junctions))
    series = []
    for i, junction in enumerate(junctions):
        g = velocity_profile[velocity_profile["junction_id"] == junction].sort_values("hour")
        series.append(
            {
                "label": f"J{junction}",
                "x": g["hour"].tolist(),
                "y": g["avg_google_speed_kmph"].tolist(),
                "color": colors[i],
                "width": 2,
            }
        )
    avg_ff = velocity_profile.groupby("junction_id")["free_flow_speed_kmph"].first().mean()
    draw_line_plot(
        series,
        PLOTS_DIR / "velocity_profile_all_junctions.png",
        "24-Hour Speed Profile - All Junctions",
        "Hour of day",
        "Speed (kmph)",
        hlines=[{"y": avg_ff, "label": f"Avg free flow {avg_ff:.1f}", "color": (220, 0, 0)}],
    )
    for i, junction in enumerate(junctions):
        g = velocity_profile[velocity_profile["junction_id"] == junction].sort_values("hour")
        ff = g["free_flow_speed_kmph"].iloc[0]
        draw_line_plot(
            [
                {
                    "label": f"Junction {junction}",
                    "x": g["hour"].tolist(),
                    "y": g["avg_google_speed_kmph"].tolist(),
                    "color": colors[i],
                    "width": 3,
                }
            ],
            PLOTS_DIR / f"velocity_profile_junction_{junction}.png",
            f"Velocity Profile - Junction {junction}",
            "Hour of day",
            "Speed (kmph)",
            hlines=[{"y": ff, "label": f"Free flow {ff:.1f}", "color": (220, 0, 0)}],
        )
    draw_bar_chart(
        speed_reduction_df["junction_id"].tolist(),
        speed_reduction_df["morning_speed_reduction_pct"].fillna(0).tolist(),
        speed_reduction_df["evening_speed_reduction_pct"].fillna(0).tolist(),
        PLOTS_DIR / "speed_reduction_bar_chart.png",
        "Peak-Hour Speed Reduction by Junction",
    )
    cong_series = []
    for i, junction in enumerate(junctions):
        g = congestion_hourly[congestion_hourly["junction_id"] == junction].sort_values("hour")
        cong_series.append(
            {
                "label": f"J{junction}",
                "x": g["hour"].tolist(),
                "y": g["avg_congestion_index"].tolist(),
                "color": colors[i],
                "width": 2,
            }
        )
    draw_line_plot(
        cong_series,
        PLOTS_DIR / "congestion_index_24hr_profile.png",
        "24-Hour Congestion Index Profile - All Junctions",
        "Hour of day",
        "Congestion index",
        hlines=[{"y": 1.0, "label": "Normal baseline", "color": (220, 0, 0)}],
    )
    draw_scatter_plot(
        delay_df["congestion_index"],
        delay_df["delay_minutes"],
        PLOTS_DIR / "congestion_vs_delay_scatter.png",
        "Congestion Index vs Delay",
        "Congestion index",
        "Delay (minutes)",
    )
    for date, g in delay_df.groupby("date"):
        pivot = g.pivot_table(index="junction_id", columns="hour", values="duration_s", aggfunc="mean")
        draw_heatmap(pivot, PLOTS_DIR / f"heatmap_{date}.png", f"Duration Heatmap - {date}")
    weekly_pivot = delay_df.pivot_table(index="junction_id", columns="hour", values="duration_s", aggfunc="mean")
    draw_heatmap(weekly_pivot, PLOTS_DIR / "heatmap_weekly_average.png", "Weekly Average Duration Heatmap")


def write_csvs(outputs):
    for name, df in outputs.items():
        path = OUTPUT_DIR / name
        df.to_csv(path, index=False)
        print(f"Saved {path.name}: {df.shape[0]} rows x {df.shape[1]} columns")


def format_excel(path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            max_len = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells[:80]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 32)
        max_row = ws.max_row
        for col_idx in range(1, ws.max_column + 1):
            values = [ws.cell(row=row, column=col_idx).value for row in range(2, min(max_row, 80) + 1)]
            numeric_values = [v for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)]
            if numeric_values and max_row >= 3:
                letter = ws.cell(row=1, column=col_idx).column_letter
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="63BE7B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="F8696B",
                    ),
                )
    wb.save(path)


def make_excel_report(outputs):
    report_path = OUTPUT_DIR / "TRAFFIC_ANALYSIS_FINAL_REPORT.xlsx"
    sheets = {
        "Junction Basic Stats": outputs["junction_basic_stats.csv"],
        "Free Flow Speed": outputs["free_flow_speed.csv"],
        "Daily Average Delay": outputs["daily_avg_delay.csv"],
        "Hourly Delay Profile": outputs["hourly_avg_delay_profile.csv"],
        "Peak Hours Per Junction": outputs["peak_hours_per_junction.csv"],
        "Speed Reduction Peak": outputs["speed_reduction_peak_hours.csv"],
        "Congestion Analysis": outputs["congestion_analysis.csv"],
        "Weekly Day Comparison": outputs["weekly_daywise_comparison.csv"],
        "Full Merged Hourly Data": outputs["full_merged_hourly_data.csv"],
    }
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet[:31], index=False)
    format_excel(report_path)
    return report_path


def professor_summary_text(summary):
    return f"""# Traffic Analysis Methodology and Results

## Data Used

The analysis uses Google Routes API traffic observations for 23 junctions from 2026-05-25 to 2026-05-31. Each junction has 24 hourly observations per day, giving:

23 junctions x 24 hours x 7 days = 3864 hourly records.

## Core Calculations

Travel duration is measured in seconds. Speed is measured in kmph.

Delay based on free-flow condition:

Delay_seconds = Actual_Duration_seconds - Free_Flow_Min_Duration_seconds

Delay_minutes = Delay_seconds / 60

Free-flow speed, primary method:

Free_Flow_Speed_kmph = Distance_km / (Minimum_Duration_seconds / 3600)

If Google route distance was available, that distance was used. Otherwise the default 10 km route length was used.

Congestion index from Google-derived summaries:

Congestion_Index = Traffic_Duration_seconds / Static_Duration_seconds

A value above 1 indicates traffic is slower than the static baseline.

## Main Weekly Findings

Worst network hour: {summary['worst_network_hour']}

Worst network congestion index: {summary['worst_network_congestion']:.4f}

Overall average free-flow speed: {summary['avg_free_flow_speed']:.2f} kmph

Overall average morning peak delay: {summary['avg_morning_delay']:.2f} minutes

Overall average evening peak delay: {summary['avg_evening_delay']:.2f} minutes

Worst junction overall: Junction {summary['worst_junction']}

Least congested junction overall: Junction {summary['best_junction']}
"""


def final_summary(delay_df, freeflow, speed_reduction_df, weekly_hourly):
    avg_delay_by_junction = (
        delay_df.groupby("junction_id", as_index=False)
        .agg(avg_delay_minutes=("delay_minutes", "mean"), avg_congestion_index=("congestion_index", "mean"))
        .sort_values("avg_delay_minutes", ascending=False)
    )
    top_delay = avg_delay_by_junction.head(5)
    top_reduction = speed_reduction_df.sort_values("max_peak_speed_reduction_pct", ascending=False).head(5)
    worst = avg_delay_by_junction.iloc[0]
    best = avg_delay_by_junction.iloc[-1]
    worst_hour = weekly_hourly.loc[weekly_hourly["avg_congestion_index"].idxmax()]
    summary = {
        "worst_network_hour": f"{int(worst_hour['hour']):02d}:00",
        "worst_network_congestion": worst_hour["avg_congestion_index"],
        "avg_free_flow_speed": freeflow["primary_free_flow_speed_kmph"].mean(),
        "avg_morning_delay": speed_reduction_df["morning_peak_delay_minutes"].mean(),
        "avg_evening_delay": speed_reduction_df["evening_peak_delay_minutes"].mean(),
        "worst_junction": worst["junction_id"],
        "best_junction": best["junction_id"],
    }
    lines = []
    lines.append("FINAL CONSOLE SUMMARY")
    lines.append("=====================")
    lines.append("\nTop 5 most congested junctions by average delay (minutes):")
    lines.append(top_delay[["junction_id", "avg_delay_minutes", "avg_congestion_index"]].round(4).to_string(index=False))
    lines.append("\nTop 5 junctions by peak-hour speed reduction (%):")
    lines.append(
        top_reduction[
            ["junction_id", "max_peak_speed_reduction_pct", "morning_speed_reduction_pct", "evening_speed_reduction_pct"]
        ].round(4).to_string(index=False)
    )
    lines.append(f"\nWorst single junction overall: Junction {summary['worst_junction']}")
    lines.append(f"Best least congested junction overall: Junction {summary['best_junction']}")
    lines.append(f"Overall average free-flow speed: {summary['avg_free_flow_speed']:.2f} kmph")
    lines.append(f"Overall average morning peak delay: {summary['avg_morning_delay']:.2f} minutes")
    lines.append(f"Overall average evening peak delay: {summary['avg_evening_delay']:.2f} minutes")
    lines.append(f"Hour with worst congestion across network: {summary['worst_network_hour']}")
    text = "\n".join(lines)
    print(text)
    (OUTPUT_DIR / "final_console_summary.txt").write_text(text, encoding="utf-8")
    (OUTPUT_DIR / "professor_explanation.md").write_text(professor_summary_text(summary), encoding="utf-8")
    return summary


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PLOTS_DIR.mkdir(parents=True, exist_ok=True)

    progress("STEP 0 - SETUP & DATA EXPLORATION")
    ensure_optional_plot_packages()
    files = file_inventory()
    exploration = explore_files(files)
    print("\nClean exploration summary:")
    print(exploration[["file", "rows", "columns", "detected_junction", "detected_hour", "detected_date", "detected_duration_s", "detected_congestion_index", "detected_speed_kmph"]].to_string(index=False))

    progress("STEP 1 - LOAD & MERGE ALL DATA")
    master = load_daily_hourly_data()
    peak_master = load_peak_data()
    weekly_peak_crosscheck = load_weekly_peak_crosscheck()
    raw = load_raw_data()
    master = add_distance(master, raw)
    print(f"Hourly records loaded: {len(master)}")
    print(f"Peak records loaded: {len(peak_master)}")
    print(f"Raw records loaded: {len(raw)}")
    print(f"Junctions present: {master['junction_id'].nunique()} -> {sorted(master['junction_id'].unique(), key=lambda x: float(x))}")
    print(f"Dates present: {sorted(master['date'].unique())}")
    print(f"Expected complete hourly rows: 23 x 7 x 24 = {23 * 7 * 24}")

    progress("STEP 2 - BASIC STATISTICS PER JUNCTION")
    stats = basic_stats(master)

    progress("STEP 3 - FREE FLOW SPEED PER JUNCTION")
    freeflow = free_flow_speed(master)

    progress("STEP 4 - DELAY CALCULATION")
    delay_df = delay_data(master, freeflow)
    daily_delay = daily_average_delay(delay_df)
    hourly_delay = hourly_delay_profile(delay_df)

    progress("STEP 5 - VELOCITY PROFILE")
    full_merged, velocity_profile = velocity_profiles(delay_df)

    progress("STEP 6 - PEAK HOUR IDENTIFICATION")
    peaks = peak_hours(delay_df, peak_master, weekly_peak_crosscheck)

    progress("STEP 7 - SPEED REDUCTION DURING PEAK HOURS")
    speed_reduction_df = speed_reduction(delay_df, peaks, freeflow)

    progress("STEP 8 - CONGESTION INDEX ANALYSIS")
    congestion_df, congestion_hourly, corr = congestion_analysis(delay_df)
    print(f"Correlation between congestion index and delay: {corr:.4f}")

    progress("STEP 9 - DAY-WISE COMPARISON")
    daywise = daywise_comparison(delay_df)
    weekly_hourly = (
        delay_df.groupby("hour", as_index=False)
        .agg(
            avg_duration_s=("duration_s", "mean"),
            avg_speed_kmph=("google_speed_kmph", "mean"),
            avg_delay_minutes=("delay_minutes", "mean"),
            avg_congestion_index=("congestion_index", "mean"),
            records=("duration_s", "count"),
        )
        .round(4)
    )

    progress("WRITING CSV OUTPUTS")
    outputs = {
        "junction_basic_stats.csv": stats,
        "free_flow_speed.csv": freeflow,
        "delay_all_hours.csv": delay_df.round(4),
        "daily_avg_delay.csv": daily_delay,
        "hourly_avg_delay_profile.csv": hourly_delay,
        "velocity_profiles.csv": velocity_profile,
        "peak_hours_per_junction.csv": peaks,
        "speed_reduction_peak_hours.csv": speed_reduction_df,
        "congestion_analysis.csv": congestion_df,
        "weekly_daywise_comparison.csv": daywise,
        "weekly_hourly_network_profile.csv": weekly_hourly,
        "full_merged_hourly_data.csv": full_merged.round(4),
        "raw_useful_columns_summary.csv": pd.DataFrame(
            [
                {
                    "raw_records": len(raw),
                    "detected_distance_column": detect_column(raw.columns, "distance_m") if not raw.empty else "",
                    "detected_duration_column": detect_column(raw.columns, "duration_s") if not raw.empty else "",
                    "detected_speed_column": detect_column(raw.columns, "speed_kmph") if not raw.empty else "",
                    "distance_used": "Google route_distance_m where available, otherwise 10 km",
                }
            ]
        ),
    }
    write_csvs(outputs)

    progress("CREATING PLOTS")
    make_plots(delay_df, velocity_profile, speed_reduction_df, congestion_hourly)
    print(f"Saved plots to {PLOTS_DIR}")

    progress("STEP 10 - MASTER EXCEL REPORT")
    report_path = make_excel_report(outputs)
    print(f"Saved Excel report: {report_path}")

    progress("STEP 11 - FINAL CONSOLE SUMMARY")
    final_summary(delay_df, freeflow, speed_reduction_df, weekly_hourly)


if __name__ == "__main__":
    main()
